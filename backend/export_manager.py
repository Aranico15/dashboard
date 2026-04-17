"""
export_manager.py - Exportación a Excel y CSV
"""
import csv
import tempfile
from datetime import date
from pathlib import Path

from sqlalchemy.orm import Session

from .database import AttendanceSession, AttendanceRecord, GroupStudent


class ExportManager:

    def _get_records(self, db: Session, session_id=None, grupo_id=None):
        q = db.query(AttendanceRecord).join(AttendanceSession)
        if session_id:
            q = q.filter(AttendanceRecord.sesion_id == session_id)
        elif grupo_id:
            q = q.filter(AttendanceSession.grupo_id == grupo_id)
        return q.order_by(AttendanceSession.fecha, AttendanceRecord.hora).all()

    def export_excel(self, db: Session, session_id=None, grupo_id=None, save_path=None) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        records = self._get_records(db, session_id, grupo_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Encabezado
        headers = ["Fecha", "Grupo", "Tipo", "Código Estudiante",
                   "Nombre", "Apellido", "Hora", "Tipo Marca", "Estado", "Confianza"]
        header_fill  = PatternFill(fill_type="solid", fgColor="4F46E5")
        header_font  = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill    = header_fill
            cell.font    = header_font
            cell.alignment = Alignment(horizontal="center")

        # Filas
        estado_colors = {"presente": "D1FAE5", "ausente": "FEE2E2", "tardanza": "FEF3C7"}
        for row_idx, r in enumerate(records, 2):
            s = r.sesion
            est = r.estudiante
            row = [
                s.fecha if s else "",
                s.grupo.nombre if s and s.grupo else "",
                s.tipo if s else "",
                est.codigo_estudiante if est else "",
                est.nombre if est else "",
                est.apellido if est else "",
                r.hora.strftime("%H:%M:%S") if r.hora else "",
                r.tipo_marca or "",
                r.estado or "",
                f"{r.confianza:.1f}%" if r.confianza else "",
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if r.estado in estado_colors:
                    cell.fill = PatternFill(fill_type="solid", fgColor=estado_colors[r.estado])

        # Ajustar ancho
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        if save_path:
            wb.save(save_path)
            return save_path
            
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        return tmp.name

    def export_csv(self, db: Session, session_id=None, grupo_id=None) -> str:
        records = self._get_records(db, session_id, grupo_id)
        tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv",
                                          delete=False, newline="", encoding="utf-8-sig")
        writer = csv.writer(tmp)
        writer.writerow(["Fecha", "Grupo", "Tipo", "Codigo", "Nombre",
                         "Apellido", "Hora", "Tipo_Marca", "Estado", "Confianza"])
        for r in records:
            s   = r.sesion
            est = r.estudiante
            writer.writerow([
                s.fecha if s else "",
                s.grupo.nombre if s and s.grupo else "",
                s.tipo if s else "",
                est.codigo_estudiante if est else "",
                est.nombre if est else "",
                est.apellido if est else "",
                r.hora.strftime("%H:%M:%S") if r.hora else "",
                r.tipo_marca or "",
                r.estado or "",
                f"{r.confianza:.1f}" if r.confianza else "",
            ])
        tmp.close()
        return tmp.name
